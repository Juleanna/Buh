import io
import zipfile
from decimal import Decimal, InvalidOperation
from datetime import datetime

import qrcode
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser
from rest_framework import status

from apps.assets.models import Asset, AssetGroup, Location


# ---------------------------------------------------------------------------
#  QR Code generation
# ---------------------------------------------------------------------------

class AssetQRCodeView(APIView):
    """
    GET /api/assets/items/{id}/qr/

    Generates a QR code PNG image for a single asset.
    The QR code encodes: inventory number, name, group, location,
    and responsible person.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            asset = Asset.objects.select_related(
                'group', 'responsible_person', 'location'
            ).get(pk=pk)
        except Asset.DoesNotExist:
            return Response(
                {'detail': 'Основний засіб не знайдено.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        qr_data = self._build_qr_data(asset)
        img = qrcode.make(qr_data)

        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)

        response = HttpResponse(buf.getvalue(), content_type='image/png')
        response['Content-Disposition'] = (
            f'inline; filename="qr_{asset.inventory_number}.png"'
        )
        return response

    @staticmethod
    def _build_qr_data(asset):
        responsible = asset.responsible_person.full_name if asset.responsible_person else ''
        location = asset.location.name if asset.location else ''

        lines = [
            f"Інв.номер: {asset.inventory_number}",
            f"Назва: {asset.name}",
            f"Група: {asset.group}",
            f"Місцезнаходження: {location}",
            f"МВО: {responsible}",
        ]
        return '\n'.join(lines)


class BulkQRCodesView(APIView):
    """
    POST /api/assets/items/bulk-qr/

    Accepts JSON body: {"asset_ids": [1, 2, 3]}
    Returns a ZIP archive containing a QR code PNG for each requested asset.
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        asset_ids = request.data.get('asset_ids', [])
        if not asset_ids or not isinstance(asset_ids, list):
            return Response(
                {'detail': 'Надайте список asset_ids.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        assets = Asset.objects.select_related(
            'group', 'responsible_person'
        ).filter(pk__in=asset_ids)

        if not assets.exists():
            return Response(
                {'detail': 'Жодного основного засобу не знайдено.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for asset in assets:
                qr_data = AssetQRCodeView._build_qr_data(asset)
                img = qrcode.make(qr_data)

                img_buf = io.BytesIO()
                img.save(img_buf, format='PNG')
                img_buf.seek(0)

                filename = f"qr_{asset.inventory_number}.png"
                zf.writestr(filename, img_buf.getvalue())

        zip_buf.seek(0)

        response = HttpResponse(zip_buf.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="qr_codes.zip"'
        return response


# ---------------------------------------------------------------------------
#  Excel export
# ---------------------------------------------------------------------------

class AssetExcelExportView(APIView):
    """
    GET /api/assets/items/export/

    Exports all active assets to an Excel (.xlsx) file with Ukrainian headers,
    proper formatting, and auto-width columns.
    """

    permission_classes = [IsAuthenticated]

    HEADERS = [
        'Інв.номер',
        'Назва',
        'Група',
        'Первісна вартість',
        'Залишкова вартість',
        'Знос',
        'Метод амортизації',
        'Дата введення',
        'МВО',
        'Місцезнаходження',
        'Статус',
    ]

    DEPRECIATION_METHOD_LABELS = dict(Asset.DepreciationMethod.choices)
    STATUS_LABELS = dict(Asset.Status.choices)

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Основні засоби'

        # ----- header row styling -----
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(
            start_color='4472C4', end_color='4472C4', fill_type='solid'
        )
        header_alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # ----- data rows -----
        data_font = Font(name='Calibri', size=11)
        data_alignment = Alignment(vertical='center', wrap_text=True)

        assets = (
            Asset.objects
            .filter(status=Asset.Status.ACTIVE)
            .select_related('group', 'responsible_person', 'location')
            .order_by('inventory_number')
        )

        for row_idx, asset in enumerate(assets, start=2):
            responsible = asset.responsible_person.full_name if asset.responsible_person else ''
            location = asset.location.name if asset.location else ''

            row_data = [
                asset.inventory_number,
                asset.name,
                str(asset.group),
                asset.initial_cost,
                asset.current_book_value,
                asset.accumulated_depreciation,
                self.DEPRECIATION_METHOD_LABELS.get(
                    asset.depreciation_method, asset.depreciation_method
                ),
                asset.commissioning_date,
                responsible,
                location,
                self.STATUS_LABELS.get(asset.status, asset.status),
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

                # Number formatting for monetary columns
                if col_idx in (4, 5, 6):
                    cell.number_format = '#,##0.00'
                # Date formatting
                if col_idx == 8 and value is not None:
                    cell.number_format = 'DD.MM.YYYY'

        # ----- auto-width -----
        for col_idx in range(1, len(self.HEADERS) + 1):
            max_length = len(str(self.HEADERS[col_idx - 1]))
            for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    if cell.value is not None:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
            adjusted_width = min(max_length + 4, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Freeze top row
        ws.freeze_panes = 'A2'

        # ----- response -----
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        content_type = (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response = HttpResponse(buf.getvalue(), content_type=content_type)
        response['Content-Disposition'] = (
            'attachment; filename="assets_export.xlsx"'
        )
        return response


# ---------------------------------------------------------------------------
#  Excel import
# ---------------------------------------------------------------------------

class AssetExcelImportView(APIView):
    """
    POST /api/assets/items/import/

    Accepts an uploaded Excel (.xlsx) file via multipart form-data.
    Expected columns (first row is header, data starts at row 2):

        Інв.номер | Назва | Код групи | Первісна вартість |
        Ліквідаційна вартість | Метод амортизації | Строк (міс.) |
        Дата введення | Місцезнаходження

    Returns JSON summary: {"created": N, "errors": [...]}.
    """

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    # Mapping of Ukrainian depreciation method labels to model choices
    METHOD_MAP = {
        'прямолінійний': Asset.DepreciationMethod.STRAIGHT_LINE,
        'зменшення залишкової вартості': Asset.DepreciationMethod.REDUCING_BALANCE,
        'прискореного зменшення залишкової вартості': Asset.DepreciationMethod.ACCELERATED_REDUCING,
        'кумулятивний': Asset.DepreciationMethod.CUMULATIVE,
        'виробничий': Asset.DepreciationMethod.PRODUCTION,
        # Also accept English keys directly
        'straight_line': Asset.DepreciationMethod.STRAIGHT_LINE,
        'reducing_balance': Asset.DepreciationMethod.REDUCING_BALANCE,
        'accelerated_reducing': Asset.DepreciationMethod.ACCELERATED_REDUCING,
        'cumulative': Asset.DepreciationMethod.CUMULATIVE,
        'production': Asset.DepreciationMethod.PRODUCTION,
    }

    EXPECTED_COLUMNS = [
        'Інв.номер',
        'Назва',
        'Код групи',
        'Первісна вартість',
        'Ліквідаційна вартість',
        'Метод амортизації',
        'Строк (міс.)',
        'Дата введення',
        'Місцезнаходження',
    ]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response(
                {'detail': 'Файл не надано. Завантажте Excel-файл у полі "file".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
        except Exception:
            return Response(
                {'detail': 'Не вдалося прочитати файл. Переконайтесь, що це .xlsx.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active
        errors = []
        created = 0

        # Pre-load groups for quick lookup
        groups_by_code = {g.code: g for g in AssetGroup.objects.all()}

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            # Skip completely empty rows
            if not row or all(cell is None for cell in row):
                continue

            # Pad the row to the expected number of columns
            row = list(row) + [None] * (len(self.EXPECTED_COLUMNS) - len(row))

            (
                inv_number,
                name,
                group_code,
                initial_cost,
                residual_value,
                depr_method,
                useful_life,
                commissioning_date,
                location,
            ) = row[:9]

            row_errors = []

            # --- Validate required fields ---
            if not inv_number:
                row_errors.append('Інвентарний номер обов\'язковий')
            else:
                inv_number = str(inv_number).strip()
                if Asset.objects.filter(inventory_number=inv_number).exists():
                    row_errors.append(
                        f'Інвентарний номер "{inv_number}" вже існує'
                    )

            if not name:
                row_errors.append('Назва обов\'язкова')
            else:
                name = str(name).strip()

            # --- Group ---
            group = None
            if not group_code:
                row_errors.append('Код групи обов\'язковий')
            else:
                group_code = str(group_code).strip()
                group = groups_by_code.get(group_code)
                if group is None:
                    row_errors.append(
                        f'Групу з кодом "{group_code}" не знайдено'
                    )

            # --- Initial cost ---
            try:
                initial_cost = Decimal(str(initial_cost))
                if initial_cost <= 0:
                    row_errors.append('Первісна вартість повинна бути > 0')
            except (InvalidOperation, TypeError, ValueError):
                row_errors.append('Невірний формат первісної вартості')
                initial_cost = None

            # --- Residual value ---
            try:
                residual_value = (
                    Decimal(str(residual_value))
                    if residual_value is not None
                    else Decimal('0.00')
                )
                if residual_value < 0:
                    row_errors.append(
                        'Ліквідаційна вартість не може бути від\'ємною'
                    )
            except (InvalidOperation, TypeError, ValueError):
                row_errors.append('Невірний формат ліквідаційної вартості')
                residual_value = Decimal('0.00')

            # --- Depreciation method ---
            method_value = Asset.DepreciationMethod.STRAIGHT_LINE
            if depr_method:
                key = str(depr_method).strip().lower()
                method_value = self.METHOD_MAP.get(key)
                if method_value is None:
                    row_errors.append(
                        f'Невідомий метод амортизації: "{depr_method}"'
                    )
                    method_value = Asset.DepreciationMethod.STRAIGHT_LINE

            # --- Useful life ---
            try:
                useful_life = int(useful_life)
                if useful_life <= 0:
                    row_errors.append(
                        'Строк корисного використання повинен бути > 0'
                    )
            except (TypeError, ValueError):
                row_errors.append(
                    'Невірний формат строку корисного використання'
                )
                useful_life = None

            # --- Commissioning date ---
            if isinstance(commissioning_date, datetime):
                commissioning_date = commissioning_date.date()
            elif isinstance(commissioning_date, str):
                for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
                    try:
                        commissioning_date = datetime.strptime(
                            commissioning_date.strip(), fmt
                        ).date()
                        break
                    except ValueError:
                        continue
                else:
                    row_errors.append(
                        f'Невірний формат дати введення: "{commissioning_date}"'
                    )
                    commissioning_date = None
            elif commissioning_date is None:
                row_errors.append('Дата введення обов\'язкова')
            # else: already a date object (openpyxl may return datetime.date)

            # --- Location ---
            location_obj = None
            if location:
                location_str = str(location).strip()
                if location_str:
                    location_obj, _ = Location.objects.get_or_create(
                        name=location_str
                    )

            # --- If errors, record and skip ---
            if row_errors:
                errors.append({'row': row_idx, 'errors': row_errors})
                continue

            # --- Compute depreciation_start_date (month after commissioning) ---
            if commissioning_date.month == 12:
                depr_start = commissioning_date.replace(
                    year=commissioning_date.year + 1, month=1, day=1
                )
            else:
                depr_start = commissioning_date.replace(
                    month=commissioning_date.month + 1, day=1
                )

            # --- Create asset ---
            try:
                Asset.objects.create(
                    inventory_number=inv_number,
                    name=name,
                    group=group,
                    initial_cost=initial_cost,
                    residual_value=residual_value,
                    current_book_value=initial_cost,
                    accumulated_depreciation=Decimal('0.00'),
                    depreciation_method=method_value,
                    useful_life_months=useful_life,
                    commissioning_date=commissioning_date,
                    depreciation_start_date=depr_start,
                    location=location_obj,
                    status=Asset.Status.ACTIVE,
                    created_by=request.user,
                )
                created += 1
            except Exception as exc:
                errors.append({'row': row_idx, 'errors': [str(exc)]})

        return Response(
            {'created': created, 'errors': errors},
            status=status.HTTP_200_OK,
        )
