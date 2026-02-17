"""
Shared openpyxl helpers for Excel export of documents.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_FONT = Font(name='Calibri', size=11)
DATA_ALIGNMENT = Alignment(vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
MONEY_FORMAT = '#,##0.00'
DATE_FORMAT = 'DD.MM.YYYY'

SECTION_FONT = Font(name='Calibri', bold=True, size=11)
TOTAL_FONT = Font(name='Calibri', bold=True, size=11)
TOTAL_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')


def create_workbook(sheet_title):
    """Create a workbook with a single sheet, return (wb, ws)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet name max 31 chars
    return wb, ws


def write_header_row(ws, row_num, headers):
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def write_data_row(ws, row_num, values, money_cols=None, date_cols=None):
    """Write a data row with optional formatting for money/date columns."""
    money_cols = money_cols or set()
    date_cols = date_cols or set()
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        if col_idx in money_cols:
            cell.number_format = MONEY_FORMAT
        if col_idx in date_cols and value is not None:
            cell.number_format = DATE_FORMAT


def write_total_row(ws, row_num, values, money_cols=None):
    """Write a totals row with bold font and light background."""
    money_cols = money_cols or set()
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER
        if col_idx in money_cols:
            cell.number_format = MONEY_FORMAT


def write_info_row(ws, row_num, label, value, label_col=1, value_col=2):
    """Write a key-value info row (e.g. 'Назва:', 'Принтер HP')."""
    label_cell = ws.cell(row=row_num, column=label_col, value=label)
    label_cell.font = Font(name='Calibri', bold=True, size=11)
    label_cell.alignment = DATA_ALIGNMENT

    value_cell = ws.cell(row=row_num, column=value_col, value=value)
    value_cell.font = DATA_FONT
    value_cell.alignment = DATA_ALIGNMENT


def write_section_header(ws, row_num, title, num_cols):
    """Write a section header spanning multiple columns."""
    cell = ws.cell(row=row_num, column=1, value=title)
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')
    if num_cols > 1:
        ws.merge_cells(
            start_row=row_num, start_column=1,
            end_row=row_num, end_column=num_cols,
        )


def auto_width(ws, num_cols, max_width=60):
    """Auto-fit column widths based on content."""
    for col_idx in range(1, num_cols + 1):
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max_length + 4, max_width,
        )


TITLE_FONT = Font(name='Calibri', bold=True, size=14)
SUBTITLE_FONT = Font(name='Calibri', bold=True, size=12)
SMALL_FONT = Font(name='Calibri', size=9, color='555555')
STAMP_FONT = Font(name='Calibri', size=9, italic=True)
SIGN_FONT = Font(name='Calibri', size=10)
SIGN_UNDERLINE_BORDER = Border(bottom=Side(style='thin'))


def write_form_header(ws, row_num, org, form_number, form_title, num_cols,
                      approval_text='Наказ Мiнiстерства фiнансiв України\n13.09.2016 №818'):
    """
    Write an official Minfin form header block matching the PDF header.
    Returns the next available row number.
    """
    org_name = org.name if org else '________________________________'
    org_edrpou = org.edrpou if org else '________'
    org_address = getattr(org, 'address', '') or ''

    # Row 1: Org name (left) + Approval stamp (right)
    org_cell = ws.cell(row=row_num, column=1, value=org_name)
    org_cell.font = Font(name='Calibri', bold=True, size=11)
    if num_cols > 1:
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=max(num_cols // 2, 2))

    stamp_text = f'Типова форма №{form_number}\nЗАТВЕРДЖЕНО\n{approval_text}'
    stamp_col = max(num_cols // 2 + 1, 3)
    stamp_cell = ws.cell(row=row_num, column=stamp_col, value=stamp_text)
    stamp_cell.font = STAMP_FONT
    stamp_cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
    if stamp_col < num_cols:
        ws.merge_cells(start_row=row_num, start_column=stamp_col,
                       end_row=row_num, end_column=num_cols)
    row_num += 1

    # Row 2: org address (if any)
    if org_address:
        addr_cell = ws.cell(row=row_num, column=1, value=org_address)
        addr_cell.font = SMALL_FONT
        row_num += 1

    # Row 3: EDRPOU
    edr_cell = ws.cell(row=row_num, column=1, value=f'Код ЄДРПОУ  {org_edrpou}')
    edr_cell.font = SMALL_FONT
    row_num += 1

    # Blank row
    row_num += 1

    # Form title
    title_cell = ws.cell(row=row_num, column=1, value=form_title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    if num_cols > 1:
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=num_cols)
    row_num += 1

    return row_num


def write_form_header_landscape(ws, row_num, org, form_title, num_cols,
                                approval_text='Наказ Мiнiстерства фiнансiв України\n13.09.2016 №818',
                                subtitle=None):
    """
    Write a landscape-oriented form header (no form number).
    Returns the next available row number.
    """
    org_name = org.name if org else '________________________________'
    org_edrpou = org.edrpou if org else '________'

    # Row 1: org name + edrpou (left) + stamp (right)
    org_cell = ws.cell(row=row_num, column=1, value=f'{org_name}   Код ЄДРПОУ  {org_edrpou}')
    org_cell.font = Font(name='Calibri', bold=True, size=11)
    mid = max(num_cols // 2, 2)
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=mid)

    stamp_text = f'ЗАТВЕРДЖЕНО\n{approval_text}'
    stamp_col = mid + 1
    stamp_cell = ws.cell(row=row_num, column=stamp_col, value=stamp_text)
    stamp_cell.font = STAMP_FONT
    stamp_cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
    if stamp_col < num_cols:
        ws.merge_cells(start_row=row_num, start_column=stamp_col,
                       end_row=row_num, end_column=num_cols)
    row_num += 2  # blank row

    # Title
    title_cell = ws.cell(row=row_num, column=1, value=form_title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=num_cols)
    row_num += 1

    # Subtitle
    if subtitle:
        sub_cell = ws.cell(row=row_num, column=1, value=subtitle)
        sub_cell.font = SUBTITLE_FONT
        sub_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=num_cols)
        row_num += 1

    row_num += 1  # blank row
    return row_num


def write_approval_block(ws, row_num, director_name='', num_cols=2):
    """
    Write 'ЗАТВЕРДЖУЮ' block with director name. Returns next row.
    """
    start_col = max(num_cols - 2, 1)
    cell = ws.cell(row=row_num, column=start_col, value='ЗАТВЕРДЖУЮ')
    cell.font = Font(name='Calibri', bold=True, size=11)
    cell.alignment = Alignment(horizontal='right')
    row_num += 1

    director_text = f'Керівник  ________  {director_name}' if director_name else 'Керівник  ________________'
    cell2 = ws.cell(row=row_num, column=start_col, value=director_text)
    cell2.font = SIGN_FONT
    cell2.alignment = Alignment(horizontal='right')
    row_num += 1

    cell3 = ws.cell(row=row_num, column=start_col, value='"___"____________ 20__ р.')
    cell3.font = SIGN_FONT
    cell3.alignment = Alignment(horizontal='right')
    row_num += 2
    return row_num


def write_text_row(ws, row_num, text, num_cols=1, font=None):
    """Write a text string merged across num_cols columns. Returns next row."""
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font = font or DATA_FONT
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    if num_cols > 1:
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=num_cols)
    return row_num + 1


def write_signatures_block(ws, row_num, signatures, num_cols=2):
    """
    Write signature lines.
    signatures: list of (role, name) tuples, e.g. [('Головний бухгалтер', 'Іваненко')]
    Returns next row.
    """
    row_num += 1  # blank row
    for role, name in signatures:
        text = f'{role}   _______________   {name}'
        cell = ws.cell(row=row_num, column=1, value=text)
        cell.font = SIGN_FONT
        if num_cols > 1:
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=min(num_cols, 6))
        row_num += 1
    return row_num


def write_commission_signatures(ws, row_num, head_name='', member_names=None, num_cols=6):
    """
    Write commission signatures block.
    Returns next row.
    """
    member_names = member_names or []
    row_num += 1

    cell = ws.cell(row=row_num, column=1,
                   value=f'Голова комісії   _______________   {head_name}')
    cell.font = SIGN_FONT
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=min(num_cols, 6))
    row_num += 1

    for i, name in enumerate(member_names):
        label = 'Члени комісії' if i == 0 else '             '
        cell = ws.cell(row=row_num, column=1,
                       value=f'{label}   _______________   {name}')
        cell.font = SIGN_FONT
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=min(num_cols, 6))
        row_num += 1

    # Blank members slots if less than 3
    remaining = max(3 - len(member_names), 0)
    for i in range(remaining):
        label = 'Члени комісії' if (i == 0 and not member_names) else '             '
        cell = ws.cell(row=row_num, column=1,
                       value=f'{label}   _______________   ')
        cell.font = SIGN_FONT
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=min(num_cols, 6))
        row_num += 1

    return row_num


def write_merged_header(ws, start_row, header_spec):
    """
    Write a complex multi-level header with merged cells.

    header_spec: list of dicts with keys:
      - text: header text
      - row: row offset (0-based)
      - col: column (1-based)
      - merge_rows: number of rows to merge (default 1)
      - merge_cols: number of cols to merge (default 1)

    All cells get HEADER_FONT, HEADER_FILL, HEADER_ALIGNMENT, THIN_BORDER.
    Returns next row after the header block.
    """
    max_row = 0
    for spec in header_spec:
        r = start_row + spec.get('row', 0)
        c = spec['col']
        mr = spec.get('merge_rows', 1)
        mc = spec.get('merge_cols', 1)

        cell = ws.cell(row=r, column=c, value=spec['text'])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

        if mr > 1 or mc > 1:
            ws.merge_cells(start_row=r, start_column=c,
                           end_row=r + mr - 1, end_column=c + mc - 1)
            # Apply border/fill to all merged cells
            for ri in range(r, r + mr):
                for ci in range(c, c + mc):
                    mcell = ws.cell(row=ri, column=ci)
                    mcell.border = THIN_BORDER
                    mcell.fill = HEADER_FILL

        max_row = max(max_row, r + mr - 1)

    return max_row + 1


def write_column_numbers_row(ws, row_num, num_cols):
    """Write a row with column numbers 1, 2, 3, ... styled like sub-header."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col_idx, value=col_idx)
        cell.font = Font(name='Calibri', size=8, italic=True, color='FFFFFF')
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    return row_num + 1


def workbook_to_response(wb, filename):
    """Save workbook to an HttpResponse."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
